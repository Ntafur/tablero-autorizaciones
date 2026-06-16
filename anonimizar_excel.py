#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Anonimiza el Excel de origen para poder publicarlo en un repositorio público.

Qué hace:
  * ELIMINA la columna `encusu` (documento de identidad del usuario).
  * ELIMINA la columna `Fecha` (campo libre no usado por el tablero).
  * SEUDONIMIZA la columna `hc` (historia clínica) con un código irreversible
    y estable (misma historia -> mismo código), para conservar el conteo de
    pacientes únicos sin exponer el número real.
  * Conserva el resto de columnas tal cual (incluidos los nombres de usuario,
    necesarios para el análisis de productividad).

El "salt" del hash es aleatorio en cada ejecución y NO se almacena, por lo que
los códigos no pueden revertirse al número de historia clínica original.

Uso:
    python3 anonimizar_excel.py [entrada.xlsx] [salida.xlsx]
"""
import sys
import os
import hashlib
import secrets
import datetime

import openpyxl

ENTRADA = sys.argv[1] if len(sys.argv) > 1 else "datos_facturas.xlsx"
SALIDA = sys.argv[2] if len(sys.argv) > 2 else "datos_tablero.xlsx"

COL_HC = "hc"
COLS_ELIMINAR = {"encusu", "Fecha"}

SALT = secrets.token_bytes(16)
_cache = {}


def seudonimo(hc):
    hc = "" if hc is None else str(hc).strip()
    if hc == "":
        return ""
    if hc not in _cache:
        h = hashlib.sha256(SALT + hc.encode("utf-8")).hexdigest()[:12]
        _cache[hc] = "P-" + h
    return _cache[hc]


def main():
    if not os.path.exists(ENTRADA):
        raise SystemExit("No se encontró el archivo de entrada: %s" % ENTRADA)

    wb_in = openpyxl.load_workbook(ENTRADA, read_only=True, data_only=True)
    ws_in = wb_in.worksheets[0]
    it = ws_in.iter_rows(values_only=True)
    header = list(next(it))

    # Índices de columnas a conservar (en su orden original)
    keep_idx = [i for i, h in enumerate(header)
                if (str(h).strip() if h is not None else "") not in COLS_ELIMINAR]
    new_header = [header[i] for i in keep_idx]
    try:
        hc_pos = new_header.index(COL_HC)
    except ValueError:
        hc_pos = -1

    wb_out = openpyxl.Workbook(write_only=True)
    ws_out = wb_out.create_sheet(title="Sheet")
    ws_out.append(new_header)

    n = 0
    for row in it:
        if row is None:
            continue
        vals = [row[i] for i in keep_idx]
        if hc_pos >= 0:
            vals[hc_pos] = seudonimo(vals[hc_pos])
        # Normaliza fechas a datetime para que se exporten como fecha real
        ws_out.append(vals)
        n += 1

    wb_out.save(SALIDA)
    print("Anonimizado: %s -> %s" % (ENTRADA, SALIDA))
    print("  Filas: %d | Pacientes seudonimizados: %d" % (n, len(_cache)))
    print("  Columnas eliminadas: %s" % ", ".join(sorted(COLS_ELIMINAR)))
    print("  Columnas conservadas: %s" % ", ".join(str(h) for h in new_header))


if __name__ == "__main__":
    main()
